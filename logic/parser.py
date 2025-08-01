"""
Excel file parsing logic with enhanced resource management and support for merged cells and complex headers
"""
import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional, Tuple
import logging
from pathlib import Path
import gc

class ExcelParser:
    """Handles parsing of Excel files with complex structures and proper resource management"""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.worksheet = None
        
    def __enter__(self):
        """Context manager entry"""
        try:
            # Force garbage collection before opening
            gc.collect()
            
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            self.worksheet = self.workbook.active
            return self
        except Exception as e:
            # Ensure cleanup on initialization failure
            self._cleanup()
            raise e
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit with guaranteed cleanup"""
        self._cleanup()
        
    def _cleanup(self):
        """Guaranteed cleanup method"""
        try:
            if self.workbook:
                self.workbook.close()
        except Exception as e:
            logging.warning(f"Error closing workbook for {self.file_path}: {e}")
        finally:
            self.workbook = None
            self.worksheet = None
            # Force garbage collection after cleanup
            gc.collect()
    
    def get_headers(self, start_row: int, end_row: int, max_columns: Optional[int] = None) -> Dict[str, int]:
        """
        Extracts unique, logical headers from a flexible, multi-row structure,
        returning a dictionary mapping the unique header name to its starting column index.

        Args:
            start_row: The starting row number for the header block (1-based).
            end_row: The ending row number for the header block (1-based).
            max_columns: Maximum number of columns to read.

        Returns:
            A dictionary mapping unique header names to their starting column index.
        """
        if not self.worksheet:
            raise ValueError("Worksheet not loaded")
        if start_row > end_row:
            raise ValueError(f"Header start row ({start_row}) cannot be after end row ({end_row}).")

        max_col = max_columns or self.worksheet.max_column
        merged_ranges = list(self.worksheet.merged_cells.ranges)  # Convert to list to avoid iteration issues
        raw_headers = []

        for col in range(1, max_col + 1):
            header_parts = []
            # Iterate through all rows in the defined header block
            for row in range(start_row, end_row + 1):
                cell_value = self._get_cell_value_with_merges(row, col, merged_ranges)
                if cell_value:
                    header_parts.append(cell_value)
            
            # Join the parts to form the final header name for the column
            final_header = ""
            if header_parts:
                # Use set to remove duplicates that can occur from vertical merges
                unique_parts = list(dict.fromkeys(header_parts))
                final_header = " - ".join(unique_parts)
            
            raw_headers.append(final_header.strip())

        # Create a dictionary with unique names, mapping to the first column index.
        # This correctly handles horizontally merged headers by only adding the first occurrence.
        unique_columns = {}
        for i, header in enumerate(raw_headers):
            # Only add the header if it has a non-empty name and has not been added before
            if header and header not in unique_columns:
                unique_columns[header] = i + 1
        
        return unique_columns
    
    def _get_cell_value_with_merges(self, row: int, col: int, merged_ranges) -> str:
        """Get cell value, handling merged cells"""
        try:
            cell = self.worksheet.cell(row=row, column=col)
            cell_value = cell.value
            
            if cell_value is not None:
                return str(cell_value).strip()
            
            # Check if cell is part of a merged range
            cell_coordinate = cell.coordinate
            for merged_range in merged_ranges:
                if cell_coordinate in merged_range:
                    # Get the top-left cell of the merged range
                    top_left = self.worksheet.cell(merged_range.min_row, merged_range.min_col)
                    if top_left.value is not None:
                        return str(top_left.value).strip()
            
            return ""
        except Exception as e:
            logging.warning(f"Error reading cell ({row}, {col}) from {self.file_path}: {e}")
            return ""
    
    def get_data_rows(self, header_row: int, headers: List[str]) -> List[Dict[str, Any]]:
        """
        Read data rows from Excel file
        
        Args:
            header_row: Row number containing headers
            headers: List of header names
            
        Returns:
            List of dictionaries with data
        """
        if not self.worksheet:
            raise ValueError("Worksheet not loaded")
        
        data = []
        start_row = header_row + 2
        
        for row in range(start_row, self.worksheet.max_row + 1):
            row_data = {}
            has_data = False
            
            for col, header in enumerate(headers, start=1):
                cell = self.worksheet.cell(row=row, column=col)
                value = cell.value
                
                if value is not None:
                    has_data = True
                    
                    # Preserve data types
                    if cell.data_type == 'n':  # Number
                        # Keep original number format
                        row_data[header] = value
                    elif cell.data_type == 'd':  # Date
                        row_data[header] = value
                    else:
                        row_data[header] = value
                else:
                    row_data[header] = None
            
            # Only add row if it contains data
            if has_data:
                data.append(row_data)
        
        return data
    
    def get_sheet_info(self) -> Dict[str, Any]:
        """Get information about the worksheet"""
        if not self.worksheet:
            raise ValueError("Worksheet not loaded")
        
        return {
            'sheet_name': self.worksheet.title,
            'max_row': self.worksheet.max_row,
            'max_column': self.worksheet.max_column,
            'merged_cells_count': len(self.worksheet.merged_cells.ranges),
            'has_filters': self.worksheet.auto_filter is not None
        }
    
    def detect_header_row(self, max_search_rows: int = 10) -> int:
        """
        Automatically detect the most likely header row
        
        Args:
            max_search_rows: Maximum number of rows to search
            
        Returns:
            Row number (1-based) that likely contains headers
        """
        if not self.worksheet:
            raise ValueError("Worksheet not loaded")
        
        best_row = 1
        best_score = 0
        
        for row in range(1, min(max_search_rows + 1, self.worksheet.max_row + 1)):
            score = self._calculate_header_score(row)
            if score > best_score:
                best_score = score
                best_row = row
        
        return best_row
    
    def _calculate_header_score(self, row: int) -> float:
        """Calculate likelihood that a row contains headers"""
        score = 0
        total_cells = 0
        
        for col in range(1, min(20, self.worksheet.max_column + 1)):  # Check first 20 columns
            cell = self.worksheet.cell(row=row, column=col)
            total_cells += 1
            
            if cell.value is not None:
                value = str(cell.value).strip()
                
                # Text values are more likely to be headers
                if cell.data_type == 's' and value:
                    score += 2
                    
                    # Common header keywords
                    header_keywords = ['name', 'id', 'date', 'amount', 'code', 'description', 
                                     'type', 'status', 'number', 'content', 'purpose']
                    if any(keyword in value.lower() for keyword in header_keywords):
                        score += 1
                
                # Numbers are less likely to be headers
                elif cell.data_type == 'n':
                    score -= 0.5
        
        return score / total_cells if total_cells > 0 else 0
    
    def validate_file(self) -> Tuple[bool, List[str]]:
        """
        Validate Excel file for common issues
        
        Returns:
            Tuple of (is_valid, list_of_issues)
        """
        issues = []
        
        try:
            if not self.file_path.exists():
                return False, ["File does not exist"]
            
            if not self.file_path.suffix.lower() in ['.xlsx', '.xls']:
                issues.append("File is not a supported Excel format (.xlsx or .xls)")
            
            # Try to open the file temporarily for validation
            temp_workbook = None
            try:
                temp_workbook = openpyxl.load_workbook(self.file_path, data_only=True)
                ws = temp_workbook.active
                
                if ws.max_row == 1:
                    issues.append("File appears to be empty or contains only one row")
                
                if ws.max_column == 1:
                    issues.append("File contains only one column")
                
                # Check for password protection
                if temp_workbook.security and temp_workbook.security.workbookPassword:
                    issues.append("File is password protected")
                
                # Check for very large files
                if ws.max_row > 100000:
                    issues.append("File is very large (>100k rows) - processing may be slow")
                
                if ws.max_column > 50:
                    issues.append("File has many columns (>50) - processing may be slow")
                    
            finally:
                if temp_workbook:
                    temp_workbook.close()
                # Force garbage collection after validation
                gc.collect()
        
        except Exception as e:
            issues.append(f"Error reading file: {str(e)}")
        
        return len(issues) == 0, issues
    
    def get_column_data_types(self, header_row: int) -> Dict[str, str]:
        """
        Analyze data types in each column
        
        Args:
            header_row: Row containing headers
            
        Returns:
            Dictionary mapping column names to predominant data types
        """
        if not self.worksheet:
            raise ValueError("Worksheet not loaded")
        
        headers = self.get_headers(header_row, header_row)
        column_types = {}
        
        # Sample first 100 data rows
        sample_size = min(100, self.worksheet.max_row - header_row)
        
        for col, header in enumerate(headers, start=1):
            type_counts = {'text': 0, 'number': 0, 'date': 0, 'empty': 0}
            
            for row in range(header_row + 1, header_row + 1 + sample_size):
                if row > self.worksheet.max_row:
                    break
                    
                cell = self.worksheet.cell(row=row, column=col)
                
                if cell.value is None:
                    type_counts['empty'] += 1
                elif cell.data_type == 'n':
                    type_counts['number'] += 1
                elif cell.data_type == 'd':
                    type_counts['date'] += 1
                else:
                    type_counts['text'] += 1
            
            # Determine predominant type
            predominant_type = max(type_counts, key=type_counts.get)
            column_types[header] = predominant_type
        
        return column_types

    def count_data_rows(self, header_end_row: int) -> int:
        """Efficiently counts the number of non-empty data rows."""
        if not self.worksheet:
            raise ValueError("Worksheet not loaded")

        count = 0
        start_data_row = header_end_row + 1
        
        # Iterate through rows and check if any cell has data
        for row in self.worksheet.iter_rows(min_row=start_data_row, values_only=True):
            if any(cell is not None for cell in row):
                count += 1
        return count

    def read_data_preview(self, headers: Dict[str, int], header_end_row: int, num_rows: int) -> List[Dict[str, Any]]:
        """Reads a specified number of data rows for preview."""
        if not self.worksheet:
            raise ValueError("Worksheet not loaded")

        data = []
        start_data_row = header_end_row + 1
        
        for row_index in range(start_data_row, min(start_data_row + num_rows, self.worksheet.max_row + 1)):
            row_data = {}
            has_data = False
            for header_name, col_index in headers.items():
                cell_value = self.worksheet.cell(row=row_index, column=col_index).value
                if cell_value is not None:
                    has_data = True
                row_data[header_name] = cell_value
            
            if has_data:
                data.append(row_data)
        
        return data

# Utility functions for external use with enhanced resource management
def quick_validate_excel(file_path: str) -> bool:
    """Quick validation of Excel file with proper cleanup"""
    parser = None
    try:
        parser = ExcelParser(file_path)
        with parser as p:
            is_valid, _ = p.validate_file()
            return is_valid
    except Exception:
        return False
    finally:
        if parser:
            parser._cleanup()
        # Force garbage collection
        gc.collect()

def get_excel_headers_safe(file_path: str, start_row: int, end_row: int) -> Dict[str, int]:
    """Safe function to get headers from Excel file with guaranteed cleanup"""
    parser = None
    try:
        parser = ExcelParser(file_path)
        with parser as p:
            return p.get_headers(start_row, end_row)
    except Exception as e:
        logging.error(f"Error getting headers from {file_path}: {str(e)}")
        return {}
    finally:
        if parser:
            parser._cleanup()
        # Force garbage collection
        gc.collect()

def get_excel_data_safe(file_path: str, header_row: int) -> Tuple[Dict[str, int], List[Dict[str, Any]]]:
    """Safe function to get headers and data from Excel file with guaranteed cleanup"""
    parser = None
    try:
        parser = ExcelParser(file_path)
        with parser as p:
            headers = p.get_headers(header_row, header_row)
            data = p.get_data_rows(header_row, list(headers.keys()))
            return headers, data
    except Exception as e:
        logging.error(f"Error getting data from {file_path}: {str(e)}")
        return {}, []
    finally:
        if parser:
            parser._cleanup()
        # Force garbage collection
        gc.collect()

def validate_excel_file_safe(file_path: str) -> Tuple[bool, List[str]]:
    """Safe function to validate Excel file with guaranteed cleanup"""
    parser = None
    try:
        parser = ExcelParser(file_path)
        with parser as p:
            return p.validate_file()
    except Exception as e:
        logging.error(f"Error validating {file_path}: {str(e)}")
        return False, [f"Validation error: {str(e)}"]
    finally:
        if parser:
            parser._cleanup()
        # Force garbage collection
        gc.collect()