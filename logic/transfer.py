"""
Data transfer engine for copying data between Excel files while preserving formatting.
This module implements a 'constructive' approach: it builds a new workbook from scratch
to avoid the limitations of modifying an existing file, especially with complex templates.
"""
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
import shutil
from pathlib import Path
import logging
from typing import List, Dict, Any, Optional, Callable, Set, Tuple
from openpyxl.utils import get_column_letter, range_boundaries, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.formula.translate import Translator
from openpyxl.cell.cell import MergedCell
import re
from copy import copy

# Dedicated logger for transfer debugging
transfer_logger = logging.getLogger('transfer_debug')
if not transfer_logger.handlers:
    transfer_logger.setLevel(logging.DEBUG)
    fh = logging.FileHandler('log.txt', mode='w', encoding='utf-8')
    fh.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - [%(funcName)s]: %(message)s')
    fh.setFormatter(formatter)
    transfer_logger.addHandler(fh)
    transfer_logger.propagate = False

def _sanitize_sheet_name(name: str) -> str:
    if not name:
        return "Untitled"
    name = str(name)
    name = re.sub(r'[\\/*?:[\]]', '', name)
    return name[:31]

def parse_skip_rows_string(skip_rows_str: str) -> Set[int]:
    skipped_rows = set()
    if not skip_rows_str:
        return skipped_rows
    for part in skip_rows_str.split(','):
        part = part.strip()
        if not part:
            continue
        if '-' in part:
            try:
                start, end = map(int, part.split('-'))
                if start <= end:
                    skipped_rows.update(range(start, end + 1))
            except ValueError:
                logging.warning(f"Could not parse range in skip_rows: {part}")
        else:
            try:
                skipped_rows.add(int(part))
            except ValueError:
                logging.warning(f"Could not parse number in skip_rows: {part}")
    return skipped_rows

class ExcelTransferEngine:
    def __init__(self, settings: Dict[str, Any], progress_callback: Optional[Callable[[int, str], None]] = None):
        self.source_path = Path(settings["source_file"])
        self.dest_path = Path(settings["dest_file"])
        self.source_header_end_row = settings["source_header_end_row"]
        self.dest_header_end_row = settings["dest_header_end_row"]
        self.dest_write_start_row = settings["dest_write_start_row"]
        self.dest_write_end_row = settings["dest_write_end_row"]
        self.group_by_column = settings.get("group_by_column")
        self.source_sheet_name = settings.get("source_sheet")
        self.master_sheet_name = settings.get("master_sheet")
        self.mappings = settings.get("mappings", {})
        self.source_columns = settings.get("source_columns", {})
        self.dest_columns = settings.get("dest_columns", {})
        self.single_value_mappings = settings.get("single_value_mapping", [])
        self.progress_callback = progress_callback

    def _update_progress(self, value: int, message: str):
        if self.progress_callback:
            self.progress_callback(value, message)

    def _get_writable_cell(self, worksheet, row_idx, col_idx):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        if isinstance(cell, MergedCell):
            for merged_range in worksheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
        return cell

    def run_transfer(self):
        transfer_logger.info("--- Starting new transfer process (Constructive Method) ---")
        if not self.group_by_column:
            raise ValueError("'Group by Column' must be selected for this operation.")
        if not self.master_sheet_name:
            raise ValueError("'Master Sheet' must be selected for this operation.")

        self._update_progress(5, "Reading source data...")
        source_data = self._read_source_data()
        if not source_data:
            raise ValueError("No data found in source file.")

        self._update_progress(15, "Grouping data...")
        grouped_data = self._group_data(source_data)
        
        self._update_progress(20, "Loading master template...")
        wb_template_vals = openpyxl.load_workbook(self.dest_path, data_only=True)
        wb_template_formulas = openpyxl.load_workbook(self.dest_path, data_only=False)

        try:
            if self.master_sheet_name not in wb_template_vals.sheetnames:
                raise ValueError(f"Master sheet '{self.master_sheet_name}' not found.")

            master_sheet_vals = wb_template_vals[self.master_sheet_name]
            master_sheet_formulas = wb_template_formulas[self.master_sheet_name]

            classified_dvs = self._classify_data_validations(master_sheet_formulas)

            output_wb = Workbook()
            first_sheet_created = False

            total_groups = len(grouped_data)
            for i, (group_name, data_rows) in enumerate(grouped_data.items()):
                self._update_progress(25 + int((i / total_groups) * 70), f"Processing group {i+1}/{total_groups}: {group_name}")
                
                new_sheet_name = _sanitize_sheet_name(group_name)
                if new_sheet_name in output_wb.sheetnames:
                    new_sheet_name = _sanitize_sheet_name(f"{group_name}_{i+1}")
                
                if not first_sheet_created:
                    # Reuse the default first sheet for the first group
                    new_sheet = output_wb.active
                    new_sheet.title = new_sheet_name
                    first_sheet_created = True
                else:
                    # Create new sheets for subsequent groups
                    new_sheet = output_wb.create_sheet(title=new_sheet_name)

                self._copy_column_dimensions(master_sheet_formulas, new_sheet)

                last_header_row = self._copy_range(master_sheet_vals, master_sheet_formulas, new_sheet, 1, self.dest_header_end_row, 1)
                last_data_row = self._write_data_rows(new_sheet, master_sheet_vals, master_sheet_formulas, data_rows, last_header_row + 1)
                
                # Determine the new start row for the footer based on user's logic
                num_template_rows = self.dest_write_end_row - self.dest_write_start_row + 1
                num_data_rows = len(data_rows)

                if num_data_rows >= num_template_rows:
                    # Expansion or same size: footer follows data
                    new_footer_start_row = last_data_row + 1
                else:
                    # Contraction: footer stays at its original relative position, after the full template space
                    new_footer_start_row = last_header_row + 1 + num_template_rows

                footer_start_row_in_master = self.dest_write_end_row + 1 if self.dest_write_end_row > 0 else 0
                if footer_start_row_in_master > 0:
                    self._copy_range(master_sheet_vals, master_sheet_formulas, new_sheet, footer_start_row_in_master, master_sheet_vals.max_row, new_footer_start_row)

                self._write_group_identifier(new_sheet, group_name, self.group_by_column)
                if data_rows:
                    self._write_single_values(new_sheet, data_rows[0])

                self._apply_classified_validations(new_sheet, classified_dvs, len(data_rows), new_footer_start_row)

            output_filename = self.dest_path.with_name(f"{self.dest_path.stem}-output{self.dest_path.suffix}")
            # Set the active sheet to the first sheet before saving to ensure workbook integrity
            if output_wb.sheetnames:
                output_wb.active = 0
            output_wb.save(output_filename)
            transfer_logger.info(f"Successfully created new file: {output_filename}")

        finally:
            wb_template_vals.close()
            wb_template_formulas.close()
            self._update_progress(100, "Transfer completed successfully")

    def _copy_column_dimensions(self, source_sheet: Worksheet, dest_sheet: Worksheet):
        """Copies all column dimension properties from a source to a destination sheet safely."""
        transfer_logger.info(f"Copying column dimensions from '{source_sheet.title}' to '{dest_sheet.title}'")
        for col_letter, dim in source_sheet.column_dimensions.items():
            new_dim = dest_sheet.column_dimensions[col_letter]
            # Manually copy all relevant properties from the source dimension.
            # This is safer than using copy.copy() which can carry over internal worksheet references.
            new_dim.min = dim.min
            new_dim.max = dim.max
            if dim.width is not None:
                new_dim.width = dim.width
            new_dim.hidden = dim.hidden
            new_dim.outline_level = dim.outline_level
            new_dim.collapsed = dim.collapsed
            # Note: dim.style is intentionally not copied. It's a reference to a style
            # in the old workbook's style table. Copying it directly causes file corruption.
            # A more complex style transfer would be needed to replicate it safely.

    def _copy_range(self, source_sheet_vals: Worksheet, source_sheet_formulas: Worksheet, dest_sheet: Worksheet, min_row: int, max_row: int, dest_start_row: int) -> int:
        transfer_logger.info(f"Copying range from {source_sheet_vals.title}:{min_row}-{max_row} to {dest_sheet.title}:{dest_start_row}")
        
        max_col = source_sheet_vals.max_column

        for r_idx, row in enumerate(source_sheet_vals.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col)):
            dest_row_idx = dest_start_row + r_idx
            if row[0].row in source_sheet_vals.row_dimensions:
                dest_sheet.row_dimensions[dest_row_idx] = copy(source_sheet_vals.row_dimensions[row[0].row])

            for c_idx, source_cell in enumerate(row):
                dest_cell = dest_sheet.cell(row=dest_row_idx, column=c_idx + 1)
                if source_cell.has_style:
                    dest_cell.font = copy(source_cell.font)
                    dest_cell.fill = copy(source_cell.fill)
                    dest_cell.number_format = source_cell.number_format
                    dest_cell.protection = copy(source_cell.protection)
                    dest_cell.alignment = copy(source_cell.alignment)
                    border_to_copy = source_cell.border
                    if (border_to_copy.left.style or border_to_copy.right.style or 
                        border_to_copy.top.style or border_to_copy.bottom.style or 
                        border_to_copy.diagonal.style):
                        dest_cell.border = copy(border_to_copy)
                
                formula = source_sheet_formulas.cell(row=source_cell.row, column=source_cell.column).value
                if isinstance(formula, str) and formula.startswith('='):
                    try:
                        translator = Translator(formula, origin=source_cell.coordinate)
                        dest_cell.value = translator.translate_formula(dest_cell.coordinate)
                    except Exception as e:
                        transfer_logger.warning(f"Could not translate formula '{formula}' from {source_cell.coordinate}. Writing as is. Error: {e}")
                        dest_cell.value = source_cell.value
                else:
                    dest_cell.value = source_cell.value

        for mc_range in source_sheet_vals.merged_cells.ranges:
            if mc_range.min_row >= min_row and mc_range.max_row <= max_row:
                offset = dest_start_row - min_row
                new_mc_coord = f"{get_column_letter(mc_range.min_col)}{mc_range.min_row + offset}:{get_column_letter(mc_range.max_col)}{mc_range.max_row + offset}"
                dest_sheet.merge_cells(new_mc_coord)
        
        return dest_start_row + (max_row - min_row)

    def _write_data_rows(self, dest_sheet: Worksheet, master_sheet_vals: Worksheet, master_sheet_formulas: Worksheet, data_rows: List[Dict[str, Any]], start_row: int) -> int:
        transfer_logger.info(f"Writing {len(data_rows)} data rows, starting at row {start_row}")
        template_row_idx = self.dest_write_start_row
        
        # Determine how many rows to create in the data area
        num_template_rows = self.dest_write_end_row - self.dest_write_start_row + 1
        num_data_rows = len(data_rows)
        # Create at least as many rows as the template has, to preserve the full formatted area
        num_rows_to_create = max(num_data_rows, num_template_rows)
        
        transfer_logger.info(f"Creating {num_rows_to_create} rows for the data area to ensure formatting is preserved.")

        # Loop to create all necessary rows and populate them
        for i in range(num_rows_to_create):
            current_write_row = start_row + i
            
            # Step 1: Copy the template row to create a fully formatted row
            self._copy_range(master_sheet_vals, master_sheet_formulas, dest_sheet, 
                             min_row=template_row_idx, max_row=template_row_idx, 
                             dest_start_row=current_write_row)

            # Step 2: If there is data for this row, overwrite the values
            if i < num_data_rows:
                row_data = data_rows[i]
                for source_col, dest_col in self.mappings.items():
                    if dest_col in self.dest_columns:
                        source_value = row_data.get(source_col)
                        # Only write to the destination cell if the source value is not empty
                        if source_value is not None and source_value != '':
                            dest_col_num = self.dest_columns[dest_col]
                            cell_to_write = self._get_writable_cell(dest_sheet, current_write_row, dest_col_num)
                            cell_to_write.value = source_value
        
        # The function must return the last row that contains *actual data*
        if num_data_rows == 0:
            return start_row - 1
        else:
            return start_row + num_data_rows - 1

    def _read_source_data(self) -> List[Dict[str, Any]]:
        workbook = None
        try:
            workbook = openpyxl.load_workbook(self.source_path, data_only=True)
            if self.source_sheet_name and self.source_sheet_name in workbook.sheetnames:
                worksheet = workbook[self.source_sheet_name]
            else:
                worksheet = workbook.active
            
            start_data_row = self.source_header_end_row + 1
            data = []
            for row_index in range(start_data_row, worksheet.max_row + 1):
                row_data, has_data = {}, False
                for header_name, col_index in self.source_columns.items():
                    value = worksheet.cell(row=row_index, column=col_index).value
                    if isinstance(value, str):
                        value = value.strip()
                    if value is not None:
                        has_data = True
                    row_data[header_name] = value
                if has_data:
                    data.append(row_data)
            return data
        finally:
            if workbook:
                workbook.close()

    def _group_data(self, source_data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        grouped = {}
        for row in source_data:
            key = str(row.get(self.group_by_column, "Uncategorized"))
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(row)
        return grouped

    def _write_group_identifier(self, worksheet, group_name: str, group_by_header: str):
        for row in worksheet.iter_rows():
            for cell in row:
                anchor_cell = self._get_writable_cell(worksheet, cell.row, cell.column)
                cell_value = str(anchor_cell.value).strip() if anchor_cell.value is not None else ""
                if cell_value.lower() == group_by_header.lower():
                    max_col = anchor_cell.column
                    for merged_range in worksheet.merged_cells.ranges:
                        if anchor_cell.coordinate in merged_range:
                            max_col = merged_range.max_col
                            break
                    target_col = max_col + 1
                    target_row = anchor_cell.row
                    try:
                        target_cell_anchor = self._get_writable_cell(worksheet, target_row, target_col)
                        target_cell_anchor.value = group_name
                        return
                    except Exception as e:
                        transfer_logger.error(f"Error writing group identifier for '{group_name}': {e}")
                        return

    def _write_single_values(self, worksheet, group_first_row: Dict[str, Any]):
        if not self.single_value_mappings:
            return
        
        transfer_logger.info(f"Writing single values to sheet '{worksheet.title}'")
        for mapping in self.single_value_mappings:
            source_col = mapping.get("source_col")
            dest_cell_addr = mapping.get("dest_cell")
            
            if not source_col or not dest_cell_addr:
                continue

            value = group_first_row.get(source_col)
            
            try:
                col_str, row_idx = coordinate_from_string(dest_cell_addr)
                col_idx = column_index_from_string(col_str)

                target_cell = self._get_writable_cell(worksheet, row_idx, col_idx)
                target_cell.value = value
                transfer_logger.debug(f"Wrote single value from '{source_col}' to {dest_cell_addr} (Value: {value})")
            except Exception as e:
                transfer_logger.error(f"Failed to write single value from '{source_col}' to {dest_cell_addr}. Error: {e}")

    def _classify_data_validations(self, master_sheet: Worksheet) -> Tuple[List, List, List, List]:
        """Scans and classifies all data validation rules into zones (header, data, footer, other)."""
        transfer_logger.debug("--- Classifying Data Validations ---")
        transfer_logger.debug(f"Settings: dest_write_start_row={self.dest_write_start_row}, dest_write_end_row={self.dest_write_end_row}")
        
        header_dvs, data_dvs, footer_dvs, other_dvs = [], [], [], []
        if not master_sheet.data_validations or not master_sheet.data_validations.dataValidation:
            transfer_logger.debug("No data validations found in master sheet.")
            return header_dvs, data_dvs, footer_dvs, other_dvs

        all_dvs = list(master_sheet.data_validations.dataValidation)
        transfer_logger.debug(f"Found {len(all_dvs)} total DV rules to classify.")

        for i, dv in enumerate(all_dvs):
            classification = "Unclassified"
            try:
                sqref_str = str(dv.sqref)
                transfer_logger.debug(f"Processing DV #{i+1}: sqref='{sqref_str}', type='{dv.type}', formula1='{dv.formula1}'")
                
                if not sqref_str:
                    transfer_logger.warning(f"DV #{i+1} has an empty sqref. Classifying as 'other'.")
                    other_dvs.append(dv)
                    continue

                in_header, in_data, in_footer, in_other = False, False, False, False
                for range_str in sqref_str.split():
                    _min_col, min_row, _max_col, max_row = range_boundaries(range_str)
                    
                    if max_row < self.dest_write_start_row:
                        in_header = True
                    elif min_row >= self.dest_write_start_row and (self.dest_write_end_row == 0 or max_row <= self.dest_write_end_row):
                        in_data = True
                    elif self.dest_write_end_row > 0 and min_row > self.dest_write_end_row:
                        in_footer = True
                    else:
                        # This logic catches ranges that span across boundaries
                        in_other = True
                
                if in_data:
                    data_dvs.append(dv)
                    classification = "data"
                elif in_footer:
                    footer_dvs.append(dv)
                    classification = "footer"
                elif in_header:
                    header_dvs.append(dv)
                    classification = "header"
                elif in_other:
                    other_dvs.append(dv)
                    classification = "other"
                
                transfer_logger.debug(f" -> Classified DV #{i+1} as: {classification}")

            except Exception as e:
                transfer_logger.error(f"Failed to classify DV #{i+1} (sqref='{dv.sqref}'). Error: {e}", exc_info=True)
                other_dvs.append(dv) # Add to 'other' on failure to be safe

        transfer_logger.info(f"Classified {len(all_dvs)} DV rules: "
                             f"{len(header_dvs)} header, {len(data_dvs)} data, {len(footer_dvs)} footer, {len(other_dvs)} other.")
        return header_dvs, data_dvs, footer_dvs, other_dvs

    def _apply_classified_validations(self, new_sheet: Worksheet, classified_dvs: Tuple, actual_data_rows: int, new_footer_start_row: int):
        """Applies pre-classified and adjusted Data Validation rules, including translating formulas inside them."""
        transfer_logger.debug(f"--- Applying Data Validations to sheet '{new_sheet.title}' ---")
        transfer_logger.debug(f"Params: actual_data_rows={actual_data_rows}, new_footer_start_row={new_footer_start_row}")
        header_dvs, data_dvs, footer_dvs, other_dvs = classified_dvs
        
        try:
            # Case 1 & 4: Header and Other rules are copied as-is
            if header_dvs or other_dvs:
                transfer_logger.debug(f"Applying {len(header_dvs)} header and {len(other_dvs)} other DVs.")
                for i, dv in enumerate(header_dvs + other_dvs):
                    transfer_logger.debug(f"  Applying other/header DV #{i+1}: sqref='{dv.sqref}'")
                    new_dv = copy(dv)
                    new_sheet.add_data_validation(new_dv)

            # Case 2: Stretch rules in the data zone
            if actual_data_rows > 0 and data_dvs:
                transfer_logger.debug(f"Applying {len(data_dvs)} data DVs for {actual_data_rows} data rows.")
                for i, dv in enumerate(data_dvs):
                    transfer_logger.debug(f"  Applying data DV #{i+1}: original sqref='{dv.sqref}'")
                    new_dv = copy(dv)
                    new_sqref = ""
                    for range_str in str(dv.sqref).split():
                        min_col, min_row, max_col, _ = range_boundaries(range_str)
                        new_max_row = self.dest_write_start_row + actual_data_rows - 1
                        if new_max_row < min_row: new_max_row = min_row
                        new_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
                        new_sqref += f" {new_range}"
                        transfer_logger.debug(f"    - Stretched range '{range_str}' to '{new_range}'")
                    
                    new_dv.sqref = new_sqref.strip()
                    if new_dv.sqref:
                        transfer_logger.debug(f"  -> Adding data DV with new sqref: '{new_dv.sqref}'")
                        new_sheet.add_data_validation(new_dv)
                    else:
                        transfer_logger.warning(f"  -> Skipping data DV #{i+1} due to empty new sqref.")

            # Case 3: Offset and translate rules in the footer
            if self.dest_write_end_row > 0 and footer_dvs:
                footer_offset = new_footer_start_row - (self.dest_write_end_row + 1)
                transfer_logger.debug(f"Applying {len(footer_dvs)} footer DVs with offset {footer_offset}.")
                for i, dv in enumerate(footer_dvs):
                    transfer_logger.debug(f"  Applying footer DV #{i+1}: original sqref='{dv.sqref}', formula1='{dv.formula1}'")
                    new_dv = copy(dv)
                    new_sqref = ""

                    first_range_str = str(dv.sqref).split()[0]
                    min_col_ref, min_row_ref, _, _ = range_boundaries(first_range_str)
                    origin_coord = f"{get_column_letter(min_col_ref)}{min_row_ref}"
                    dest_coord = f"{get_column_letter(min_col_ref)}{min_row_ref + footer_offset}"
                    transfer_logger.debug(f"    - Translation coords: origin='{origin_coord}', dest='{dest_coord}'")

                    # Translate formulas if they are actual range references
                    if dv.formula1 and isinstance(dv.formula1, str) and dv.formula1.startswith('='):
                        try:
                            translator = Translator(dv.formula1, origin=origin_coord)
                            new_dv.formula1 = translator.translate_formula(dest_coord)
                            transfer_logger.debug(f"    - Translated formula1 to: '{new_dv.formula1}'")
                        except Exception as e:
                            transfer_logger.warning(f"    - Could not translate DV formula1 '{dv.formula1}'. Using original. Error: {e}")
                    
                    if dv.formula2 and isinstance(dv.formula2, str) and dv.formula2.startswith('='):
                        try:
                            translator = Translator(dv.formula2, origin=origin_coord)
                            new_dv.formula2 = translator.translate_formula(dest_coord)
                            transfer_logger.debug(f"    - Translated formula2 to: '{new_dv.formula2}'")
                        except Exception as e:
                            transfer_logger.warning(f"    - Could not translate DV formula2 '{dv.formula2}'. Using original. Error: {e}")

                    # Offset the sqref range itself
                    for range_str in str(dv.sqref).split():
                        min_col, min_row, max_col, max_row = range_boundaries(range_str)
                        new_min_row = min_row + footer_offset
                        new_max_row = max_row + footer_offset
                        new_range = f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}"
                        new_sqref += f" {new_range}"
                        transfer_logger.debug(f"    - Offset range '{range_str}' to '{new_range}'")
                    
                    new_dv.sqref = new_sqref.strip()
                    if new_dv.sqref:
                        transfer_logger.debug(f"  -> Adding footer DV with new sqref: '{new_dv.sqref}' and formula1: '{new_dv.formula1}'")
                        new_sheet.add_data_validation(new_dv)
                    else:
                        transfer_logger.warning(f"  -> Skipping footer DV #{i+1} due to empty new sqref.")

        except Exception as e:
            transfer_logger.error(f"CRITICAL FAILURE in applying DV rules for sheet '{new_sheet.title}'. Error: {e}", exc_info=True)