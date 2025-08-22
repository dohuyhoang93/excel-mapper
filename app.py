import tkinter as tk
from tkinter import ttk, filedialog, simpledialog
import ttkbootstrap as ttk_boot
from ttkbootstrap.constants import *
import json
import os
import logging
from pathlib import Path
from typing import Optional, List, Dict, Any
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import subprocess
import sys
from datetime import datetime
import traceback
import shutil
from threading import Thread
import gc
import time
import psutil
from logic.parser import ExcelParser
from logic.config_manager import ConfigurationManager
from logic.mapper import ColumnMapper
from logic.transfer import ExcelTransferEngine, parse_skip_rows_string
from gui.widgets import (ScrollableFrame, AboutDialog, PreviewDialog, 
                         DetectionConfigDialog, show_custom_info, 
                         show_custom_error, show_custom_warning, 
                         show_custom_question)

import json
from pathlib import Path

# --- Logging Configuration ---
def setup_logging():
    """Reads logging configuration and sets up the root logger."""
    try:
        config_path = Path("configs/app_settings.json")
        if config_path.exists():
            with open(config_path, 'r') as f:
                config = json.load(f)
            log_level_str = config.get("log_level", "INFO").upper()
        else:
            log_level_str = "INFO"

        log_level = getattr(logging, log_level_str, logging.INFO)

        # Configure root logger
        logging.basicConfig(
            level=log_level,
            format='%(asctime)s - %(name)s - %(levelname)s - [%(funcName)s]: %(message)s',
            filename='app.log',
            filemode='w', 
            encoding='utf-8'
        )
        logging.info("Logging configured based on settings.")
    except Exception as e:
        # Fallback basic config if setup fails
        logging.basicConfig(level=logging.INFO, filename='app.log', filemode='w', encoding='utf-8')
        logging.error("Failed to setup logging from config file.", exc_info=True)

class FileHandleManager:
    """Manages file handles to prevent Excel file locking issues"""
    
    @staticmethod
    def force_release_handles():
        """Force garbage collection and release file handles"""
        for _ in range(3):
            gc.collect()
            time.sleep(0.05)
    
    @staticmethod
    def is_file_locked(file_path: str) -> bool:
        """Check if a file is currently locked by another process"""
        try:
            with open(file_path, 'r+b'):
                return False
        except (IOError, OSError):
            return True
    
    @staticmethod
    def wait_for_file_release(file_path: str, max_wait_seconds: int = 5) -> bool:
        """Wait for a file to be released by other processes"""
        start_time = time.time()
        while time.time() - start_time < max_wait_seconds:
            if not FileHandleManager.is_file_locked(file_path):
                return True
            FileHandleManager.force_release_handles()
            time.sleep(0.2)
        return False
    
    @staticmethod
    def get_processes_using_file(file_path: str) -> list:
        """Get list of processes that are using the specified file"""
        processes = []
        try:
            for proc in psutil.process_iter(['pid', 'name', 'open_files']):
                try:
                    if proc.info['open_files']:
                        for file_info in proc.info['open_files']:
                            if os.path.samefile(file_info.path, file_path):
                                processes.append({'pid': proc.info['pid'], 'name': proc.info['name']})
                                break
                except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess, OSError):
                    continue
        except Exception as e:
            logging.warning(f"Error checking file usage: {e}")
        return processes

class ExcelDataMapper:
    def __init__(self):
        self.root = ttk_boot.Window(themename="flatly")
        self.root.title("Excel Data Mapper")
        self.root.geometry("1000x800")

        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
        self.icon_path = None
        try:
            base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
            icon_path = os.path.join(base_path, "icon.ico")
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
                self.icon_path = icon_path
        except Exception:
            pass
        
        # --- Variable Declarations ---
        self.source_file = tk.StringVar()
        self.dest_file = tk.StringVar()
        self.source_sheet = tk.StringVar()
        self.master_sheet = tk.StringVar()
        
        self.source_header_start_row = tk.IntVar(value=1)
        self.source_header_end_row = tk.IntVar(value=1)
        self.dest_header_start_row = tk.IntVar(value=9)
        self.dest_header_end_row = tk.IntVar(value=9)
        
        self.group_by_column = tk.StringVar()
        self.current_theme = "flatly"
        self.preview_limit_var = tk.IntVar(value=1000)
        
        self.dest_write_start_row = tk.IntVar(value=11)
        self.dest_write_end_row = tk.IntVar(value=0)
        self.detection_keywords = tk.StringVar(value="total,sum,cộng,tổng,thành tiền")
        
        self.source_columns = {}
        self.dest_columns = {}
        self.mapping_combos = {}

        # Dynamic list for single value mappings
        self.single_value_mappings_list: List[Dict[str, Any]] = []
        
        self.config_manager = ConfigurationManager()
        self.column_mapper = ColumnMapper()

        self.setup_menu()
        self.setup_gui()
        self.load_app_settings()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
    def setup_menu(self):
        menubar_frame = ttk_boot.Frame(self.root)
        menubar_frame.pack(fill=X, side=TOP, padx=5, pady=(1, 0))

        def create_menu_button(text, menu):
            button = ttk_boot.Button(menubar_frame, text=text, bootstyle="link")
            button.pack(side=LEFT, padx=10, pady=1)
            button.config(command=lambda: menu.post(button.winfo_rootx(), button.winfo_rooty() + button.winfo_height()))
            return button

        file_menu = tk.Menu(self.root, tearoff=0)
        file_menu.add_command(label="Open Destination Folder", command=self.open_dest_folder)
        file_menu.add_command(label="Force Release File Handles", command=self.force_release_excel_handles)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        create_menu_button("File", file_menu)

        settings_menu = tk.Menu(self.root, tearoff=0)
        settings_menu.add_command(label="Switch Theme", command=self.toggle_theme)
        settings_menu.add_command(label="Configure Detection...", command=self.open_detection_config_dialog)
        create_menu_button("Settings", settings_menu)

        about_menu = tk.Menu(self.root, tearoff=0)
        about_menu.add_command(label="Info", command=self.show_about)
        create_menu_button("About", about_menu)
        
    def setup_gui(self):
        main_frame = ttk_boot.Frame(self.root, padding=5)
        main_frame.pack(fill=BOTH, expand=True, side=TOP)

        self.status_frame = ttk_boot.Frame(main_frame)
        self.status_frame.pack(side=BOTTOM, fill=X, pady=(5, 0))
        self.status_frame.columnconfigure(0, weight=1)
        self.status_frame.columnconfigure(1, weight=0)

        self.progress = ttk_boot.Progressbar(self.status_frame, mode='determinate', bootstyle=SUCCESS)
        self.progress.grid(row=0, column=0, sticky='we', padx=(0, 5))
        
        self.status_label = ttk_boot.Label(self.status_frame, text="Ready")
        self.status_label.grid(row=0, column=1, sticky='e')

        action_frame = ttk_boot.Frame(main_frame)
        action_frame.pack(side=BOTTOM, fill=X, pady=5)
        self.save_button = ttk_boot.Button(action_frame, text="Save Configuration", command=self.save_config, bootstyle=SUCCESS)
        self.save_button.pack(side=LEFT, padx=(0, 5))
        self.load_button = ttk_boot.Button(action_frame, text="Load Configuration", command=self.load_config, bootstyle=INFO)
        self.load_button.pack(side=LEFT, padx=5)
        
        self.execute_button = ttk_boot.Button(action_frame, text="Execute Transfer", command=self.execute_transfer, bootstyle=PRIMARY)
        self.execute_button.pack(side=RIGHT, padx=0)

        preview_frame = ttk_boot.Frame(action_frame)
        preview_frame.pack(side=RIGHT, padx=5)

        self.preview_button = ttk_boot.Button(preview_frame, text="Preview Transfer", command=self.preview_transfer, bootstyle="outline-secondary")
        self.preview_button.pack(side=LEFT, padx=(0,5))

        ttk_boot.Label(preview_frame, text="Limit (rows):").pack(side=LEFT)
        self.preview_limit_spinbox = ttk_boot.Spinbox(preview_frame, from_=0, to=999999, textvariable=self.preview_limit_var, width=8)
        self.preview_limit_spinbox.pack(side=LEFT)

        content_frame = ttk_boot.Frame(main_frame)
        content_frame.pack(fill=BOTH, expand=True)

        left_panel = ttk_boot.Frame(content_frame, padding=(0, 0, 10, 0))
        left_panel.pack(side=LEFT, fill=Y, anchor=N)

        right_panel = ttk_boot.Frame(content_frame)
        right_panel.pack(side=LEFT, fill=BOTH, expand=True)

        # --- Populate Left Panel ---
        file_frame = ttk_boot.LabelFrame(left_panel, text="File & Sheet Selection", padding=5)
        file_frame.pack(fill=X, pady=(0, 5), anchor=N)
        file_frame.columnconfigure(1, weight=1)
        
        ttk_boot.Label(file_frame, text="Source File:").grid(row=0, column=0, sticky=W, pady=2)
        ttk_boot.Entry(file_frame, textvariable=self.source_file, width=50).grid(row=0, column=1, padx=5, pady=2, sticky=EW)
        ttk_boot.Button(file_frame, bootstyle="outline", text="Browse", command=self.browse_source_file).grid(row=0, column=2, padx=5, pady=2)
        
        ttk_boot.Label(file_frame, text="Source Sheet:").grid(row=1, column=0, sticky=W, pady=2)
        self.source_sheet_combo = ttk_boot.Combobox(file_frame, textvariable=self.source_sheet, state=DISABLED)
        self.source_sheet_combo.grid(row=1, column=1, columnspan=2, padx=5, pady=2, sticky=EW)

        ttk_boot.Label(file_frame, text="Destination File:").grid(row=2, column=0, sticky=W, pady=2)
        ttk_boot.Entry(file_frame, textvariable=self.dest_file, width=50).grid(row=2, column=1, padx=5, pady=2, sticky=EW)
        ttk_boot.Button(file_frame, bootstyle="outline", text="Browse", command=self.browse_dest_file).grid(row=2, column=2, padx=5, pady=2)

        ttk_boot.Label(file_frame, text="Master Sheet:").grid(row=3, column=0, sticky=W, pady=2)
        self.master_sheet_combo = ttk_boot.Combobox(file_frame, textvariable=self.master_sheet, state=DISABLED)
        self.master_sheet_combo.grid(row=3, column=1, columnspan=2, padx=5, pady=2, sticky=EW)

        header_frame = ttk_boot.LabelFrame(left_panel, text="Header Configuration", padding=5)
        header_frame.pack(fill=X, pady=(0, 5), anchor=N)
        header_frame.columnconfigure(4, weight=1)

        ttk_boot.Label(header_frame, text="Source:").grid(row=0, column=0, sticky=W, pady=(0, 2))
        ttk_boot.Label(header_frame, text="From:").grid(row=0, column=1, sticky=W, padx=(5, 0))
        ttk_boot.Spinbox(header_frame, from_=1, to=50, textvariable=self.source_header_start_row, width=5).grid(row=0, column=2, padx=(5, 10))
        ttk_boot.Label(header_frame, text="To:").grid(row=0, column=3, sticky=W)
        ttk_boot.Spinbox(header_frame, from_=1, to=50, textvariable=self.source_header_end_row, width=5).grid(row=0, column=4, padx=5, sticky=EW)
        
        ttk_boot.Label(header_frame, text="Destination:").grid(row=1, column=0, sticky=W, pady=(0, 2))
        ttk_boot.Label(header_frame, text="From:").grid(row=1, column=1, sticky=W, padx=(5, 0))
        ttk_boot.Spinbox(header_frame, from_=1, to=50, textvariable=self.dest_header_start_row, width=5).grid(row=1, column=2, padx=(5, 10))
        ttk_boot.Label(header_frame, text="To:").grid(row=1, column=3, sticky=W)
        ttk_boot.Spinbox(header_frame, from_=1, to=50, textvariable=self.dest_header_end_row, width=5).grid(row=1, column=4, padx=5, sticky=EW)
        
        self.load_cols_button = ttk_boot.Button(header_frame, text="Load Columns", command=self.safe_load_columns, bootstyle="outline-info")
        self.load_cols_button.grid(row=0, column=5, rowspan=2, padx=(10, 0), pady=2, sticky="ns")

        write_zone_frame = ttk_boot.LabelFrame(left_panel, text="Setting write zone", padding=5)
        write_zone_frame.pack(fill=X, pady=(0, 5), anchor=N)
        write_zone_frame.columnconfigure(1, weight=1)
        ttk_boot.Label(write_zone_frame, text="Start Write Row:").grid(row=0, column=0, sticky=W, padx=(0,5), pady=2)
        ttk_boot.Spinbox(write_zone_frame, from_=1, to=99999, textvariable=self.dest_write_start_row, width=8).grid(row=0, column=1, sticky=W)
        ttk_boot.Label(write_zone_frame, text="End Write Row:").grid(row=1, column=0, sticky=W, padx=(0,5), pady=2)
        ttk_boot.Spinbox(write_zone_frame, from_=0, to=99999, textvariable=self.dest_write_end_row, width=8).grid(row=1, column=1, sticky=W)
        self.detect_button = ttk_boot.Button(write_zone_frame, text="Detect Zone", command=self.detect_write_zone, bootstyle="outline-info")
        self.detect_button.grid(row=0, column=2, rowspan=2, padx=(5,0), pady=2, sticky="ns")

        group_by_frame = ttk_boot.LabelFrame(left_panel, text="Group by Configuration", padding=5)
        group_by_frame.pack(fill=X, pady=(0, 5), anchor=N)
        group_by_frame.columnconfigure(0, weight=1)
        ttk_boot.Label(group_by_frame, text="Group by Column:").pack(fill=X)
        self.group_by_combo = ttk_boot.Combobox(group_by_frame, textvariable=self.group_by_column)
        self.group_by_combo.pack(fill=X)

        # --- Dynamic Single Value Mapping Frame ---
        single_value_lf = ttk_boot.LabelFrame(left_panel, text="Single Value Mapping", padding=5)
        single_value_lf.pack(fill=X, pady=(0, 5), anchor=N)
        single_value_lf.rowconfigure(0, weight=1)
        single_value_lf.columnconfigure(0, weight=1)

        self.single_value_scroll_frame = ScrollableFrame(single_value_lf, height=150)
        self.single_value_scroll_frame.grid(row=0, column=0, sticky="nsew")

        # --- Right Panel ---
        mapping_container = ttk_boot.LabelFrame(right_panel, text="Column Mapping", padding=5)
        mapping_container.pack(fill=BOTH, expand=True)
        self.mapping_scroll_frame = ScrollableFrame(mapping_container)
        self.mapping_scroll_frame.pack(fill=BOTH, expand=True)

        # Initial draw
        self._redraw_single_value_frame()

    def add_single_value_row(self):
        """Adds a new empty row to the single value mapping list and redraws the UI."""
        self.single_value_mappings_list.append({
            "source_var": tk.StringVar(),
            "dest_var": tk.StringVar()
        })
        self._redraw_single_value_frame()

    def remove_single_value_row(self, row_to_remove: Dict[str, Any]):
        """Removes a specific row from the single value mapping list and redraws the UI."""
        self.single_value_mappings_list.remove(row_to_remove)
        self._redraw_single_value_frame()

    def _redraw_single_value_frame(self):
        """Clears and redraws all widgets in the single value mapping frame."""
        for widget in self.single_value_scroll_frame.scrollable_frame.winfo_children():
            widget.destroy()

        container = self.single_value_scroll_frame.scrollable_frame
        container.columnconfigure(0, weight=1)
        container.columnconfigure(1, weight=1)
        container.columnconfigure(2, weight=0)
        container.columnconfigure(3, weight=0)

        # --- Draw Headers ---
        ttk_boot.Label(container, text="Source Column", font='-weight bold').grid(row=0, column=0, sticky="w", padx=2)
        ttk_boot.Label(container, text="Destination Cell", font='-weight bold').grid(row=0, column=1, sticky="w", padx=2)

        # If list is empty, add one blank row to start
        if not self.single_value_mappings_list:
            self.single_value_mappings_list.append({"source_var": tk.StringVar(), "dest_var": tk.StringVar()})

        source_keys = list(self.source_columns.keys())

        # --- Draw Mapping Rows ---
        for i, mapping_item in enumerate(self.single_value_mappings_list, start=1):
            combo = ttk_boot.Combobox(container, textvariable=mapping_item["source_var"], values=source_keys)
            combo.grid(row=i, column=0, sticky="ew", padx=(0, 2), pady=2)
            
            entry = ttk_boot.Entry(container, textvariable=mapping_item["dest_var"], width=15)
            entry.grid(row=i, column=1, sticky="ew", padx=2, pady=2)

            remove_btn = ttk_boot.Button(container, text="-", bootstyle="danger-outline", width=3,
                                         command=lambda item=mapping_item: self.remove_single_value_row(item))
            remove_btn.grid(row=i, column=2, padx=(2,0), pady=2)
            
            # Add the "+" button only to the first row
            if i == 1:
                add_btn = ttk_boot.Button(container, text="+", bootstyle="success-outline", width=3,
                                          command=self.add_single_value_row)
                add_btn.grid(row=i, column=3, padx=(2,0), pady=2)
    
    def browse_source_file(self):
        filename = filedialog.askopenfilename(title="Select Source Excel file", filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            self.source_file.set(filename)
            self.log_info(f"Source file selected: {filename}")
            self._load_source_sheets(filename)
    
    def browse_dest_file(self):
        filename = filedialog.askopenfilename(title="Select Destination Excel file", filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            self.dest_file.set(filename)
            self.log_info(f"Destination file selected: {filename}")
            self._load_destination_sheets(filename)

    def _load_source_sheets(self, file_path: str):
        """Loads VISIBLE sheet names from the source file into the source sheet combobox."""
        try:
            if not file_path or not os.path.exists(file_path):
                self.source_sheet_combo['values'] = []
                self.source_sheet.set('')
                self.source_sheet_combo.config(state=DISABLED)
                return

            FileHandleManager.force_release_handles()
            wb = openpyxl.load_workbook(file_path, read_only=True)
            visible_sheets = [sheet.title for sheet in wb.worksheets if sheet.sheet_state == 'visible']
            wb.close()
            
            self.source_sheet_combo['values'] = visible_sheets
            if visible_sheets:
                self.source_sheet.set(visible_sheets[0])
                self.source_sheet_combo.config(state='readonly')
            else:
                self.source_sheet.set('')
                self.source_sheet_combo.config(state=DISABLED)
            self.log_info(f"Loaded visible source sheets: {visible_sheets}")

        except Exception as e:
            self.log_error(f"Error loading source sheets: {str(e)}")
            show_custom_error(self.root, self, "Error", f"Could not read sheets from source file: {str(e)}")
            self.source_sheet_combo['values'] = []
            self.source_sheet.set('')
            self.source_sheet_combo.config(state=DISABLED)

    def _load_destination_sheets(self, file_path: str):
        """Loads VISIBLE sheet names from the destination file into the master sheet combobox."""
        try:
            if not file_path or not os.path.exists(file_path):
                self.master_sheet_combo['values'] = []
                self.master_sheet.set('')
                self.master_sheet_combo.config(state=DISABLED)
                return

            FileHandleManager.force_release_handles()
            wb = openpyxl.load_workbook(file_path, read_only=True)
            visible_sheets = [sheet.title for sheet in wb.worksheets if sheet.sheet_state == 'visible']
            wb.close()
            
            self.master_sheet_combo['values'] = visible_sheets
            if visible_sheets:
                self.master_sheet.set(visible_sheets[0])
                self.master_sheet_combo.config(state='readonly')
            else:
                self.master_sheet.set('')
                self.master_sheet_combo.config(state=DISABLED)
            self.log_info(f"Loaded visible destination sheets: {visible_sheets}")

        except Exception as e:
            self.log_error(f"Error loading destination sheets: {str(e)}")
            show_custom_error(self.root, self, "Error", f"Could not read sheets from destination file: {str(e)}")
            self.master_sheet_combo['values'] = []
            self.master_sheet.set('')
            self.master_sheet_combo.config(state=DISABLED)
    
    def force_release_excel_handles(self):
        try:
            self.source_columns, self.dest_columns = {}, {}
            FileHandleManager.force_release_handles()
            self.update_status("File handles released")
            self.log_info("Forced release of Excel file handles")
            show_custom_info(self.root, self, "Info", "Excel file handles have been released.")
        except Exception as e:
            self.log_error(f"Error forcing handle release: {str(e)}")
            show_custom_error(self.root, self, "Error", f"Error releasing handles: {str(e)}")
    
    def check_file_accessibility(self, file_path: str) -> bool:
        try:
            if not os.path.exists(file_path):
                show_custom_error(self.root, self, "File Not Found", f"The file could not be found:\n{file_path}")
                return False
            if FileHandleManager.is_file_locked(file_path):
                self.update_status(f"Waiting for file to be released: {os.path.basename(file_path)}")
                if not FileHandleManager.wait_for_file_release(file_path, max_wait_seconds=10):
                    processes = FileHandleManager.get_processes_using_file(file_path)
                    if processes:
                        process_names = [p['name'] for p in processes]
                        self.log_error(f"File locked by processes: {', '.join(process_names)}")
                        show_custom_warning(self.root, self, "File Locked", f"File is locked by: {', '.join(process_names)}\nPlease close these applications and try again.")
                    return False
            return True
        except Exception as e:
            self.log_error(f"Error checking file accessibility: {str(e)}")
            return False
    
    def get_excel_columns(self, file_path, start_row, end_row, sheet_name: Optional[str] = None):
        try:
            FileHandleManager.force_release_handles()
            with ExcelParser(file_path, sheet_name=sheet_name) as parser:
                if not parser.worksheet:
                    raise ValueError(f"Sheet '{sheet_name}' not found or workbook is empty.")
                headers = parser.get_headers(start_row, end_row)
                return {name: index for name, index in headers.items() if name and str(name).strip()}
        except Exception as e:
            self.log_error(f"Error reading Excel columns with parser: {str(e)}")
            raise
        finally:
            FileHandleManager.force_release_handles()

    def safe_load_columns(self, saved_group_by_col: Optional[str] = None, apply_suggestions: bool = True):
        try:
            if not self.source_file.get() or not self.dest_file.get():
                show_custom_warning(self.root, self, "Warning", "Please select both source and destination files first.")
                return
            if not self.check_file_accessibility(self.source_file.get()) or not self.check_file_accessibility(self.dest_file.get()):
                return
            
            self.force_release_excel_handles()
            self.update_status("Loading columns...")
            
            self.source_columns = self.get_excel_columns(self.source_file.get(), self.source_header_start_row.get(), self.source_header_end_row.get(), sheet_name=self.source_sheet.get())
            self.dest_columns = self.get_excel_columns(self.dest_file.get(), self.dest_header_start_row.get(), self.dest_header_end_row.get(), sheet_name=self.master_sheet.get())
            
            if not self.source_columns or not self.dest_columns:
                show_custom_error(self.root, self, "Error", "Could not load columns. Please check file paths, sheet selections, and header row numbers.")
                return
            
            source_keys = list(self.source_columns.keys())
            self.group_by_combo['values'] = source_keys
            if saved_group_by_col and saved_group_by_col in source_keys:
                self.group_by_column.set(saved_group_by_col)

            # Update comboboxes for single value mappings
            self._redraw_single_value_frame()
            
            self.create_mapping_widgets(apply_suggestions=apply_suggestions)
            self.update_status(f"Loaded {len(self.source_columns)} source and {len(self.dest_columns)} destination columns")
            self.log_info("Columns loaded successfully")
        except Exception as e:
            self.log_error(f"Error loading columns: {str(e)}")
            show_custom_error(self.root, self, "Error", f"Failed to load columns: {str(e)}")
            self.update_status("Error loading columns")
        finally:
            FileHandleManager.force_release_handles()
    
    def create_mapping_widgets(self, apply_suggestions: bool = True):
        for widget in self.mapping_scroll_frame.scrollable_frame.winfo_children():
            widget.destroy()
        
        self.mapping_scroll_frame.scrollable_frame.columnconfigure(0, weight=1)
        self.mapping_scroll_frame.scrollable_frame.columnconfigure(2, weight=1)
        ttk_boot.Label(self.mapping_scroll_frame.scrollable_frame, text="Source Column", font="-weight bold").grid(row=0, column=0, sticky=W, padx=5, pady=(2, 2))
        ttk_boot.Label(self.mapping_scroll_frame.scrollable_frame, text="Destination Column", font="-weight bold").grid(row=0, column=2, sticky=W, padx=5, pady=(2, 2))
        
        self.mapping_combos = {}
        for i, source_col_name in enumerate(self.source_columns.keys(), start=1):
            ttk_boot.Label(self.mapping_scroll_frame.scrollable_frame, text=source_col_name, anchor=W).grid(row=i, column=0, sticky=EW, padx=5, pady=2)
            ttk_boot.Label(self.mapping_scroll_frame.scrollable_frame, text="→").grid(row=i, column=1, sticky=W, padx=5)
            dest_combo = ttk_boot.Combobox(self.mapping_scroll_frame.scrollable_frame, values=[""] + list(self.dest_columns.keys()), width=60)
            dest_combo.grid(row=i, column=2, sticky=EW, padx=5, pady=2)
            if apply_suggestions:
                suggested = self.column_mapper.suggest_mapping(source_col_name, list(self.dest_columns.keys()))
                if suggested: dest_combo.set(suggested)
            self.mapping_combos[source_col_name] = dest_combo
    
    def save_config(self):
        try:
            if not hasattr(self, 'mapping_combos') or not self.mapping_combos:
                show_custom_warning(self.root, self, "Warning", "No mappings to save. Please load columns first.")
                return
            
            config_file_path = filedialog.asksaveasfilename(
                title="Save Job Configuration As",
                defaultextension=".json",
                filetypes=[("JSON files", "*.json")],
                initialdir=self.config_manager.config_dir
            )
            if not config_file_path: return

            mappings = {source_col: combo.get() for source_col, combo in self.mapping_combos.items() if combo.get()}
            
            single_value_mappings = [
                {"source_col": item["source_var"].get(), "dest_cell": item["dest_var"].get()}
                for item in self.single_value_mappings_list
                if item["source_var"].get() and item["dest_var"].get()
            ]

            job_config = {
                "source_header_start_row": self.source_header_start_row.get(), "source_header_end_row": self.source_header_end_row.get(),
                "dest_header_start_row": self.dest_header_start_row.get(), "dest_header_end_row": self.dest_header_end_row.get(),
                "dest_write_start_row": self.dest_write_start_row.get(), "dest_write_end_row": self.dest_write_end_row.get(),
                "group_by_column": self.group_by_column.get(), 
                "source_sheet": self.source_sheet.get(),
                "master_sheet": self.master_sheet.get(),
                "mapping": mappings,
                "single_value_mapping": single_value_mappings
            }
            
            self.config_manager.save_job_config(job_config, config_file_path)
            self.save_app_settings()
            
            self.update_status(f"Job configuration saved to {os.path.basename(config_file_path)}")
            self.log_info(f"Job configuration saved: {config_file_path}")
        except Exception as e:
            self.log_error(f"Error saving job configuration: {str(e)}")
            show_custom_error(self.root, self, "Error", f"Failed to save job configuration: {str(e)}")
    
    def load_config(self):
        try:
            config_file_path = filedialog.askopenfilename(
                title="Load Job Configuration", 
                filetypes=[("JSON files", "*.json")],
                initialdir=self.config_manager.config_dir
            )
            if not config_file_path: return

            config = self.config_manager.load_job_config(config_file_path)

            self.source_header_start_row.set(config.get("source_header_start_row", 1))
            self.source_header_end_row.set(config.get("source_header_end_row", 1))
            self.dest_header_start_row.set(config.get("dest_header_start_row", 9))
            self.dest_header_end_row.set(config.get("dest_header_end_row", 9))
            self.dest_write_start_row.set(config.get("dest_write_start_row", self.dest_header_end_row.get() + 1))
            self.dest_write_end_row.set(config.get("dest_write_end_row", 0))

            # Load and validate source sheet
            saved_source_sheet = config.get("source_sheet", "")
            available_source_sheets = self.source_sheet_combo['values']
            if available_source_sheets and saved_source_sheet and saved_source_sheet in available_source_sheets:
                self.source_sheet.set(saved_source_sheet)
            elif saved_source_sheet:
                show_custom_warning(self.root, self, "Source Sheet Not Found", 
                                    f"Saved source sheet '{saved_source_sheet}' not found in the current source file.\n"
                                    f"Defaulting to the first available sheet: '{available_source_sheets[0] if available_source_sheets else ''}'.")

            # Load and validate master sheet (destination)
            saved_master_sheet = config.get("master_sheet", "")
            available_dest_sheets = self.master_sheet_combo['values']
            if available_dest_sheets and saved_master_sheet and saved_master_sheet in available_dest_sheets:
                self.master_sheet.set(saved_master_sheet)
            elif saved_master_sheet:
                show_custom_warning(self.root, self, "Master Sheet Not Found", 
                                    f"Saved master sheet '{saved_master_sheet}' not found in the current destination file.\n"
                                    f"Defaulting to the first available sheet: '{available_dest_sheets[0] if available_dest_sheets else ''}'.")

            # Load single value mappings from config
            self.single_value_mappings_list.clear()
            loaded_sv_mappings = config.get("single_value_mapping", [])
            for item in loaded_sv_mappings:
                self.single_value_mappings_list.append({
                    "source_var": tk.StringVar(value=item.get("source_col", "")),
                    "dest_var": tk.StringVar(value=item.get("dest_cell", ""))
                })
            self._redraw_single_value_frame()

            if self.source_file.get() and self.dest_file.get():
                saved_group_by_col = config.get("group_by_column", "")
                self.safe_load_columns(saved_group_by_col=saved_group_by_col, apply_suggestions=False)
                mappings = config.get("mapping", {})
                for source_col, dest_col in mappings.items():
                    if source_col in self.mapping_combos:
                        self.mapping_combos[source_col].set(dest_col)

            self.save_app_settings()
            self.update_status(f"Job configuration loaded from {os.path.basename(config_file_path)}")
            self.log_info(f"Job configuration loaded: {config_file_path}")
        except Exception as e:
            self.log_error(f"Error in load_config: {str(e)}")
            show_custom_error(self.root, self, "Error", f"Failed to load job configuration: {str(e)}")
    
    def load_app_settings(self):
        """Loads global application settings at startup."""
        try:
            settings = self.config_manager.load_app_settings()
            
            new_theme = settings.get("theme", "flatly")
            if new_theme != self.current_theme:
                self.root.style.theme_use(new_theme)
                self.current_theme = new_theme
            
            self.detection_keywords.set(settings.get("detection_keywords", "total,sum,cộng,tổng,thành tiền"))

            last_source = settings.get("last_source_file", "")
            if os.path.exists(last_source):
                self.source_file.set(last_source)
                self._load_source_sheets(last_source)
            
            last_dest = settings.get("last_dest_file", "")
            if os.path.exists(last_dest):
                self.dest_file.set(last_dest)
                self._load_destination_sheets(last_dest)

            self.log_info("Application settings loaded.")

        except Exception as e:
            self.log_error(f"Error loading application settings: {str(e)}")
            
    def save_app_settings(self):
        """Saves global application settings, preserving existing ones."""
        try:
            # First, load the existing settings to preserve keys like 'log_level'
            try:
                settings = self.config_manager.load_app_settings()
            except FileNotFoundError:
                settings = {} # If the file doesn't exist, start with an empty dict

            # Now, update the settings with the current application state
            settings["theme"] = self.current_theme
            settings["detection_keywords"] = self.detection_keywords.get()
            settings["last_source_file"] = self.source_file.get()
            settings["last_dest_file"] = self.dest_file.get()
            
            # Ensure log_level is preserved or set a default if it's missing
            if "log_level" not in settings:
                settings["log_level"] = "INFO"

            self.config_manager.save_app_settings(settings)
            self.log_info("Application settings saved.")
        except Exception as e:
            self.log_error(f"Error saving application settings: {str(e)}")

    def on_closing(self):
        """Handles window closing event."""
        self.save_app_settings()
        self.root.destroy()
    
    def execute_transfer(self, excluded_groups: Optional[List[str]] = None):
        if not self.source_file.get() or not self.dest_file.get():
            show_custom_warning(self.root, self, "Warning", "Please select both source and destination files.")
            return
        if not self.source_sheet.get() or not self.master_sheet.get():
            show_custom_warning(self.root, self, "Warning", "Please select a sheet for both source and destination files.")
            return
        if not hasattr(self, 'mapping_combos') or not self.mapping_combos:
            show_custom_warning(self.root, self, "Warning", "Please load columns first.")
            return
        if not self.group_by_column.get():
            show_custom_warning(self.root, self, "Warning", "Please select a 'Group by Column'.")
            return
        
        mappings = {s: c.get() for s, c in self.mapping_combos.items() if c.get()}
        if not mappings:
            show_custom_warning(self.root, self, "Warning", "Please configure at least one column mapping.")
            return
        dest_values = list(mappings.values())
        duplicates = [d for d in set(dest_values) if dest_values.count(d) > 1]
        if duplicates:
            show_custom_error(self.root, self, "Error", f"Duplicate destination columns detected: {', '.join(duplicates)}")
            return
        if not self.check_file_accessibility(self.source_file.get()) or not self.check_file_accessibility(self.dest_file.get()):
            return

        self.disable_controls()
        self.update_status("Starting data transfer...")
        self.progress['value'] = 0
        transfer_thread = Thread(target=self._execute_transfer_thread, args=(mappings, excluded_groups))
        transfer_thread.daemon = True
        transfer_thread.start()

    def _execute_transfer_thread(self, mappings, excluded_groups: Optional[List[str]] = None):
        try:
            single_value_mappings = [
                {"source_col": item["source_var"].get(), "dest_cell": item["dest_var"].get()}
                for item in self.single_value_mappings_list
                if item["source_var"].get() and item["dest_var"].get()
            ]

            settings = {
                "source_file": self.source_file.get(),
                "dest_file": self.dest_file.get(),
                "source_sheet": self.source_sheet.get(),
                "master_sheet": self.master_sheet.get(),
                "source_header_start_row": self.source_header_start_row.get(),
                "source_header_end_row": self.source_header_end_row.get(),
                "dest_header_start_row": self.dest_header_start_row.get(),
                "dest_header_end_row": self.dest_header_end_row.get(),
                "dest_write_start_row": self.dest_write_start_row.get(),
                "dest_write_end_row": self.dest_write_end_row.get(),
                "group_by_column": self.group_by_column.get(),
                "mappings": mappings,
                "source_columns": self.source_columns,
                "dest_columns": self.dest_columns,
                "single_value_mapping": single_value_mappings
            }

            engine = ExcelTransferEngine(settings, self.update_progress_callback)

            # --- Data Reading and Filtering ---
            source_data = engine._read_source_data()
            if not source_data:
                raise ValueError("No data found in source file.")
            
            grouped_data = engine._group_data(source_data)

            if excluded_groups:
                self.log_info(f"Excluding {len(excluded_groups)} groups from transfer.")
                grouped_data = {k: v for k, v in grouped_data.items() if k not in excluded_groups}

            if not grouped_data:
                self.root.after(0, self.on_transfer_error, ValueError("All groups were excluded or no groups were found."))
                return

            engine.run_transfer(grouped_data=grouped_data)
            
            self.root.after(0, self.on_transfer_success)
        except Exception as e:
            self.root.after(0, self.on_transfer_error, e)
        finally:
            FileHandleManager.force_release_handles()
            self.root.after(0, self.enable_controls)

    def update_progress_callback(self, value: int, message: str):
        """Callback function for the engine to update the GUI's progress."""
        self.progress['value'] = value
        self.update_status(message)
        self.root.update()
    
    def on_transfer_success(self):
        self.progress['value'] = 100
        self.update_status("Transfer completed successfully")
        show_custom_info(self.root, self, "Success", "Data transfer completed successfully!")
        if show_custom_question(self.root, self, "Open Folder", "Would you like to open the destination folder?"):
            self.open_dest_folder()

    def on_transfer_error(self, error):
        self.log_error(f"Error during transfer thread: {str(error)}\n{traceback.format_exc()}")
        self.update_status("Transfer failed")
        self.progress['value'] = 0
        show_custom_error(self.root, self, "Error", f"Transfer failed: {str(error)}")

    def disable_controls(self):
        for widget in [self.execute_button, self.load_button, self.save_button, self.load_cols_button, self.preview_button, self.preview_limit_spinbox]:
            widget.config(state=DISABLED)

    def enable_controls(self):
        for widget in [self.execute_button, self.load_button, self.save_button, self.load_cols_button, self.preview_button, self.preview_limit_spinbox]:
            widget.config(state=NORMAL)
    
    def toggle_theme(self):
        new_theme = "superhero" if self.current_theme == "flatly" else "flatly"
        self.root.style.theme_use(new_theme)
        self.current_theme = new_theme
        self.update_status(f"Theme changed to {new_theme}")
    
    def open_dest_folder(self):
        if not self.dest_file.get():
            show_custom_warning(self.root, self, "Warning", "No destination file selected.")
            return
        dest_path = Path(self.dest_file.get())
        if dest_path.exists():
            folder_path = dest_path.parent
            try:
                if os.name == 'nt':
                    os.startfile(folder_path)
                elif sys.platform == 'darwin':
                    subprocess.run(['open', folder_path])
                else:
                    subprocess.run(['xdg-open', folder_path])
            except Exception as e:
                self.log_error(f"Could not open folder: {e}")
                show_custom_error(self.root, self, "Error", f"Could not open folder:\n{e}")
        else:
            show_custom_warning(self.root, self, "Warning", "Destination file does not exist.")
    
    def show_about(self):
        AboutDialog(self.root, self)
    
    def update_status(self, message):
        self.status_label.config(text=message)
        self.root.update_idletasks()
    
    def log_info(self, message): logging.info(message)
    def log_warning(self, message): logging.warning(message)
    def log_error(self, message): logging.error(message)

    def open_detection_config_dialog(self):
        DetectionConfigDialog(self.root, self)

    def detect_write_zone(self):
        if not self.dest_file.get() or not os.path.exists(self.dest_file.get()):
            show_custom_warning(self.root, self, "Warning", "Please select a valid destination file first.")
            return
        try:
            self.update_status("Detecting write zone...")
            predicted_start_row = self.dest_header_end_row.get() + 1
            self.dest_write_start_row.set(predicted_start_row)
            predicted_end_row = 0
            with ExcelParser(self.dest_file.get(), sheet_name=self.master_sheet.get()) as p:
                if not p.worksheet: return
                ws = p.worksheet
                keywords = [k.strip() for k in self.detection_keywords.get().lower().split(',') if k.strip()]
                for row in range(predicted_start_row, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell_val = ws.cell(row, col).value
                        if isinstance(cell_val, str) and any(k in cell_val.lower() for k in keywords):
                            predicted_end_row = row - 1; break
                    if predicted_end_row > 0: break
                if predicted_end_row == 0:
                    for row in range(predicted_start_row, ws.max_row + 1):
                        if all(ws.cell(row, col).value is None for col in range(1, ws.max_column + 1)):
                            predicted_end_row = row - 1; break
            if predicted_end_row >= predicted_start_row:
                self.dest_write_end_row.set(predicted_end_row)
                self.update_status(f"Detection complete. Start: {predicted_start_row}, End: {predicted_end_row}.")
            else:
                self.dest_write_end_row.set(0)
                self.update_status(f"Detection complete. Start: {predicted_start_row}, End: Unlimited.")
        except Exception as e:
            self.log_error(f"Error detecting write zone: {str(e)}")
            show_custom_error(self.root, self, "Error", f"Failed to detect write zone: {str(e)}")
            self.update_status("Detection failed")

    def _run_preview_simulation(self):
        report = {}
        try:
            group_by_col = self.group_by_column.get()
            if not group_by_col:
                report["error"] = "Please select a 'Group by Column' for the preview."
                return report

            # --- 1. Get Data and Settings ---
            temp_settings = {
                "source_file": self.source_file.get(),
                "dest_file": self.dest_file.get(),
                "source_sheet": self.source_sheet.get(),
                "source_header_end_row": self.source_header_end_row.get(),
                "dest_header_end_row": self.dest_header_end_row.get(),
                "dest_write_start_row": self.dest_write_start_row.get(),
                "dest_write_end_row": self.dest_write_end_row.get(),
                "source_columns": self.source_columns,
                "group_by_column": group_by_col
            }
            engine = ExcelTransferEngine(temp_settings)
            row_limit = self.preview_limit_var.get()
            source_data = engine._read_source_data(row_limit=row_limit if row_limit > 0 else None)

            if not source_data:
                report["error"] = "No data found in source file to generate a preview."
                return report

            # --- 2. Get Validation Rules from Destination ---
            validation_errors = []
            dest_validations = []
            try:
                with ExcelParser(self.dest_file.get(), sheet_name=self.master_sheet.get()) as p:
                    dest_validations = p.get_data_validations()
            except Exception as e:
                self.log_warning(f"Could not read data validations from destination: {e}")

            # --- 3. Perform Validation Simulation ---
            if dest_validations:
                mappings = {s: c.get() for s, c in self.mapping_combos.items() if c.get()}
                for i, row in enumerate(source_data):
                    for source_col, dest_col in mappings.items():
                        if not dest_col: continue
                        
                        dest_col_idx = self.dest_columns.get(dest_col)
                        if not dest_col_idx: continue

                        for dv in dest_validations:
                            if f"{openpyxl.utils.get_column_letter(dest_col_idx)}" in str(dv['sqref']):
                                if dv['type'] == 'list':
                                    source_value = row.get(source_col)
                                    if source_value is not None:
                                        allowed_values = [str(v).strip() for v in dv['formula1'].replace('"', '').split(',')]
                                        if str(source_value).strip() not in allowed_values:
                                            validation_errors.append({
                                                "row": i + self.source_header_end_row.get() + 1,
                                                "column": dest_col,
                                                "value": source_value,
                                                "rule": f"Value must be one of: {dv['formula1']}"
                                            })
                                break

            # --- 4. Finalize Report ---
            grouped_data = engine._group_data(source_data)
            report['group_count'] = len(grouped_data)
            report['total_rows'] = len(source_data)
            report['row_limit'] = row_limit
            all_groups = sorted(grouped_data.items())
            report['groups'] = [(name, len(rows)) for name, rows in all_groups]
            report['validation_errors'] = validation_errors

            return report
        except Exception as e:
            self.log_error(f"Error during preview simulation: {str(e)}")
            self.log_error(traceback.format_exc())
            report["error"] = f"An error occurred: {str(e)}"
            return report
        finally:
            FileHandleManager.force_release_handles()

    def preview_transfer(self):
        if not all([self.source_file.get(), os.path.exists(self.source_file.get()), self.dest_file.get(), os.path.exists(self.dest_file.get())]):
            show_custom_error(self.root, self, "Error", "Please select valid source and destination files.")
            return
        if not self.source_sheet.get() or not self.master_sheet.get():
            show_custom_warning(self.root, self, "Warning", "Please select a sheet for both source and destination.")
            return
        if not hasattr(self, 'mapping_combos') or not self.source_columns:
            show_custom_warning(self.root, self, "Warning", "Please load columns first.")
            return
        self.update_status("Generating simulation report...")
        try:
            report_data = self._run_preview_simulation()
            if "error" in report_data:
                PreviewDialog(self.root, self, report_data, [], {}); return
            mappings = {s: c.get() for s, c in self.mapping_combos.items() if c.get()}
            if not mappings:
                show_custom_warning(self.root, self, "Warning", "Please configure at least one mapping for a meaningful preview.")
                return
            with ExcelParser(self.source_file.get(), sheet_name=self.source_sheet.get()) as p:
                if not p.worksheet: raise ValueError(f"Sheet '{self.source_sheet.get()}' not found.")
                preview_data = p.read_data_preview(self.source_columns, self.source_header_end_row.get(), 10)
            report_data['settings'] = self.get_current_settings()
            dialog = PreviewDialog(self.root, self, report_data, preview_data, mappings)
            if dialog.result is not None:
                self.execute_transfer(excluded_groups=dialog.result)
            else:
                self.update_status("Preview closed.")
        except Exception as e:
            self.log_error(f"Error generating preview: {str(e)}\n{traceback.format_exc()}")
            show_custom_error(self.root, self, "Error", f"Failed to generate preview: {str(e)}")
            self.update_status("Preview failed")

    def get_current_settings(self) -> dict:
        return {
            "Source File": os.path.basename(self.source_file.get()), "Destination File": os.path.basename(self.dest_file.get()),
            "Group by Column": self.group_by_column.get() or "None", "---": "---",
            "Start Write Row": self.dest_write_start_row.get(), "End Write Row": self.dest_write_end_row.get() or "Unlimited",
        }
    
    def run(self):
        try:
            self.root.mainloop()
        except Exception as e:
            self.log_error(f"Critical error in main loop: {str(e)}")
            try:
                root = tk.Tk()
                root.withdraw()
                show_custom_error(None, None, "Critical Error", f"Application encountered a critical error:\n\n{e}")
            except tk.TclError:
                print(f"CRITICAL ERROR: {e}")
                input("Press Enter to exit...")

if __name__ == "__main__":
    setup_logging()
    try:
        app = ExcelDataMapper()
        app.run()
    except Exception as e:
        logging.critical(f"Failed to start application: {str(e)}\n{traceback.format_exc()}")
        try:
            root = tk.Tk()
            root.withdraw()
            from tkinter import messagebox
            messagebox.showerror("Critical Error", f"Application encountered a critical error:\n\n{e}")
        except tk.TclError:
            print(f"CRITICAL ERROR: {e}")
            input("Press Enter to exit...")